import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook
from operator import itemgetter

wb = load_workbook('data.xlsx')
ws = wb.active

storedSheet = []
newData = []

cell_range = ws['A1':'D11']#this will loop and read all data in worksheet

for row in cell_range:
	date = row[0].value
	name = row[1].value
	isbn = row[2].value
	price = row[3].value

	storedSheet = [price,date,name,isbn]
	priceFirst = storedSheet
	newData.append(priceFirst)

newData = newData[1:]
print sorted(newData, key = lambda x: float(x[0]))

